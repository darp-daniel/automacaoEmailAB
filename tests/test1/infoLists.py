from pprint import pprint
import openpyxl
import time as t
import pandas as pd
import sqlite3

def extrair_clients(file_path, linhasPc, colunaInfo):
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
    except Exception as e:
        print(f"Erro ao abrir o arquivo: {e}")
        return []
    

    clientes = []

    max_linhas = ws.max_row
    blocos =   max_linhas // linhasPc
    try:
        for bloco in range(blocos):
        
            inicio = (bloco * linhasPc) + 1
    
            info_cliente = {}
    
            for info, (col, linha) in colunaInfo.items():
                cell = ws[f"{col}{inicio + linha - 1}"]
                info_cliente[info] = str(cell.value).strip() if cell.value is not None else 'Sem informação'
            
            clientes.append(info_cliente)
        return clientes
    except Exception as e:
        print(f"Erro ao extrair clientes: {e}")
        return []

def criar_db(clientes):
    try:
        df = pd.DataFrame(clientes)
        df = df[df['email'] != 'Sem informação']
        df = df.drop_duplicates(subset=["email"], keep='first')
        df = df.reset_index(drop=True)
        df['email_enviado'] = False

        conn = sqlite3.connect('src/db/clientes_test1.db')
        df.to_sql('clientes', conn, if_exists='replace', index=False)

        print("Dados inseridos no banco de dados com sucesso!")
    except Exception as e:
        print(f"Erro ao criar o banco de dados: {e}")
    

   
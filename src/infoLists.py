from pprint import pprint
import openpyxl
import time as t
import pandas as pd
import sqlite3

def extrair_clients(file_path, linhasPc, colunaInfo):
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
    except Exception as e:
        print(f"Erro ao abrir o arquivo: {e}")
        return []
    

    clientes = []

    max_linhas = ws.max_row
    blocos =   max_linhas // linhasPc
    try:
        for bloco in range(blocos):
        
            inicio = (bloco * linhasPc) + 1
    
            info_cliente = {}
    
            for info, (col, linha) in colunaInfo.items():
                cell = ws[f"{col}{inicio + linha - 1}"]
                info_cliente[info] = str(cell.value).strip() if cell.value is not None else 'Sem informação'
            
            clientes.append(info_cliente)
        return clientes
    except Exception as e:
        print(f"Erro ao extrair clientes: {e}")
        return []

if __name__ == "__main__":
    file_path = "/home/daniel/Desktop/WorkSpace/automacaoEmails/src/Atualização cadastral.xlsx"
    linhasPc = 12
    colunaInfo1 = {
        "nomeEmpresa": ("D", 1),
        "email": ("B", 8),
        "representante": ("C", 6),
        "cargo": ("J", 6)
    }
    colunaInfo2 = {
        "nomeEmpresa": ("P", 1),
        "email": ("N", 8),
        "representante": ("O", 6),
        "cargo": ("V", 6)
    }
    clients1 = extrair_clients(file_path, linhasPc, colunaInfo1)
    clients2 = extrair_clients(file_path, linhasPc, colunaInfo2)

    df = pd.DataFrame(clients1 + clients2)
    df = df[df['email'] != 'sem informação']
    df = df.drop_duplicates(subset=["email"], keep='first')
    df = df.reset_index(drop=True)
    df['email_enviado'] = False

    conn = sqlite3.connect('src/db/clientes.db')
    df.to_sql('clientes', conn, if_exists='replace', index=False)
    t.sleep(5)

    query = """SELECT * FROM clientes WHERE email_enviado = 0
 ORDER BY RANDOM() LIMIT 15 """
    df = pd.read_sql_query(query, conn)
    conn.close()
    pprint(df.to_dict(orient='records'))

   